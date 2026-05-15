import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importamos todo lo del archivo visual y la base de datos
from config_ui import *
from database import Database

class MainWin(tk.Toplevel):
    TABS = ["Panel general", "Lista de alumnos", "Estadísticas", "Nuevo registro"]

    def __init__(self, master, db, user):
        super().__init__(master)
        self.db, self.user = db, user
        self.title("Sistema Lince")
        self.geometry("1140x720")
        self.minsize(920, 600)
        self.configure(bg=C_GRIS_BG)
        self._shell()
        self._nav("Panel general")
        self.grab_set()

    # ── Shell (header + area) ──────────────────────────────────────
    def _shell(self):
        hdr = tk.Frame(self, bg=C_VINO, height=86)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        inner = tk.Frame(hdr, bg=C_VINO)
        inner.pack(fill="both", expand=True, padx=22)

        # Izquierda: titulo + usuario
        left = tk.Frame(inner, bg=C_VINO)
        left.pack(side="left", pady=8)
        tk.Label(left, text="◈  Sistema Lince",
                 bg=C_VINO, fg=C_BLANCO,
                 font=("Georgia", 19, "bold")).pack(anchor="w")
        nombre_u = self.user.get("nombre", self.user.get("correo", "Usuario"))
        rol_u    = self.user.get("rol", "")
        tk.Label(left, text=f"Bienvenido, {nombre_u}  ·  {rol_u}",
                 bg=C_VINO, fg="#FF8A80", font=F_SMALL).pack(anchor="w")

        # Derecha: botones nav
        nav = tk.Frame(inner, bg=C_VINO)
        nav.pack(side="right", pady=18)
        ICONS = {"Panel general": "⊞", "Lista de alumnos": "☰",
                 "Estadísticas": "⊕", "Nuevo registro": "＋"}
        self._nb = {}
        for t in self.TABS:
            icon = ICONS.get(t, "")
            b = tk.Button(
                nav, text=f"  {icon} {t}  ", font=F_NAV,
                bg=C_VINO, fg=C_BLANCO, relief="flat",
                bd=0, cursor="hand2", padx=10, pady=5,
                activebackground="#5C0000",
                command=lambda tab=t: self._nav(tab),
            )
            b.pack(side="left", padx=2)
            self._nb[t] = b

        # Barra color separador
        tk.Frame(self, bg="#5C0000", height=3).pack(fill="x")

        self._area = tk.Frame(self, bg=C_GRIS_BG)
        self._area.pack(fill="both", expand=True)

    def _nav(self, tab):
        for t, b in self._nb.items():
            b.config(bg="#5C0000" if t == tab else C_VINO)
        for w in self._area.winfo_children():
            w.destroy()
        {
            "Panel general":    self._panel,
            "Lista de alumnos": self._lista,
            "Estadísticas":     self._stats,
            "Nuevo registro":   self._nuevo,
        }[tab]()

    # ══════════════════════════════════════════════════════════════
    #  PANEL GENERAL
    # ══════════════════════════════════════════════════════════════
    def _panel(self):
        st   = self.db.get_stats()
        now  = datetime.datetime.now()
        MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Panel general",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer,
                 text="Resumen de prospectos registrados vía chatbot",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Tarjetas estadisticas
        crow = tk.Frame(outer, bg=C_GRIS_BG)
        crow.pack(fill="x", pady=(0, 20))

        cards = [
            ("Total\nregistrados", str(st.get("total", 0)), C_VINO),
            ("Nuevos\n(Hoy)", str(st.get("hoy", 0)), C_AZUL_TAG),
            ("Bachilleratos\ndistintos", str(st.get("bachilleratos", 0)), C_VERDE_TAG),
            ("Carrera top\n" + st.get("carrera_top", "—"),
             f"{st.get('carrera_top_pct', 0)}%", C_MORADO_TAG),
        ]
        for lbl, val, accent in cards:
            c = tk.Frame(crow, bg=C_BLANCO,
                         highlightbackground=C_GRIS_BORDE,
                         highlightthickness=1)
            c.pack(side="left", fill="both", expand=True, padx=5)
            tk.Frame(c, bg=accent, height=4).pack(fill="x")
            tk.Label(c, text=val, bg=C_BLANCO, fg=accent,
                     font=F_STAT_N).pack(pady=(14, 4))
            tk.Label(c, text=lbl, bg=C_BLANCO, fg=C_SUBTEXTO,
                     font=F_STAT_L, justify="center").pack(pady=(0, 14))

        # Recientes
        rh = tk.Frame(outer, bg=C_GRIS_BG)
        rh.pack(fill="x", pady=(0, 8))
        tk.Label(rh, text="Registros recientes",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H3).pack(side="left")
        btn_pill(rh, "Ver todos →",
                 lambda: self._nav("Lista de alumnos"),
                 bg=C_BLANCO, fg=C_VINO, px=12, py=4).pack(side="right")

        self._tabla(outer, self.db.get_prospectos()[:10])

    # ══════════════════════════════════════════════════════════════
    #  LISTA DE ALUMNOS
    # ══════════════════════════════════════════════════════════════
    def _lista(self):
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Lista de alumnos",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer, text="Prospectos registrados desde el chatbot",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Controles
        ctrl = tk.Frame(outer, bg=C_GRIS_BG)
        ctrl.pack(fill="x", pady=(0, 10))

        opciones  = [("Todas las carreras", None)] + \
                    [(n, i) for i, n in CARRERAS_NOMBRE.items()]
        self._fmap = {o[0]: o[1] for o in opciones}
        nombres   = [o[0] for o in opciones]

        self._cv = tk.StringVar(value="Todas las carreras")
        ttk.Combobox(ctrl, textvariable=self._cv, values=nombres,
                     state="readonly", font=F_BODY, width=28).pack(
            side="left", ipady=4, padx=(0, 10))
        self._cv.trace_add("write", lambda *_: self._reload_tabla())

        btn_pill(ctrl, "⬇  Exportar Excel", self._exportar,
                 bg="#37474F", px=14, py=5).pack(side="left")

        self._clbl = tk.Label(ctrl, text="", bg=C_GRIS_BG,
                              fg=C_SUBTEXTO, font=F_SMALL)
        self._clbl.pack(side="right")

        self._touter = tk.Frame(outer, bg=C_GRIS_BG)
        self._touter.pack(fill="both", expand=True)
        self._reload_tabla()

    def _reload_tabla(self):
        for w in self._touter.winfo_children():
            w.destroy()
        fid  = self._fmap.get(self._cv.get()) if hasattr(self, "_cv") else None
        rows = self.db.get_prospectos(fid)
        self._clbl.config(text=f"{len(rows)} registros")
        self._tabla(self._touter, rows, num=True)

    # ══════════════════════════════════════════════════════════════
    #  ESTADISTICAS
    # ══════════════════════════════════════════════════════════════
    def _stats(self):
        st    = self.db.get_stats()
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Estadísticas",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer,
                 text="Bachilleratos con más registros y carreras de mayor interés",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        brow = tk.Frame(outer, bg=C_GRIS_BG)
        brow.pack(fill="x", pady=(0, 18))
        btn_pill(brow, "⬇  Reporte bachilleratos",
                 lambda: self._exp_rep("bachilleratos"),
                 bg="#37474F", px=14, py=5).pack(side="left", padx=(0, 10))
        btn_pill(brow, "⬇  Reporte carreras",
                 lambda: self._exp_rep("carreras"),
                 bg="#37474F", px=14, py=5).pack(side="left")

        gr = tk.Frame(outer, bg=C_GRIS_BG)
        gr.pack(fill="both", expand=True)

        self._barchart(
            gr, "Bachilleratos con más alumnos",
            st.get("bach_stats", []),
            "bach", "cnt", C_AZUL_TAG,
        ).pack(side="left", fill="both", expand=True, padx=(0, 12))

        self._barchart(
            gr, "Carreras con mayor interés",
            st.get("carrera_stats", []),
            "nombre_carrera", "cnt", C_VERDE_TAG,
        ).pack(side="left", fill="both", expand=True)

    def _barchart(self, parent, title, data, lkey, vkey, color):
        c = card_frame(parent)
        tk.Frame(c, bg=color, height=4).pack(fill="x")
        tk.Label(c, text=title, bg=C_BLANCO, fg=C_TEXTO,
                 font=F_H3, anchor="w").pack(anchor="w", padx=16, pady=(12, 8))
        tk.Frame(c, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        mx = max((r[vkey] for r in data), default=1) or 1
        for r in data:
            lbl = str(r.get(lkey, "—"))
            val = r.get(vkey, 0)
            rf  = tk.Frame(c, bg=C_BLANCO)
            rf.pack(fill="x", padx=16, pady=5)

            tk.Label(rf, text=lbl, bg=C_BLANCO, fg=C_TEXTO,
                     font=F_SMALL, width=20, anchor="w").pack(side="left")

            trk = tk.Frame(rf, bg="#EEEEEE", height=14, width=180)
            trk.pack(side="left", padx=6)
            trk.pack_propagate(False)
            fw = max(2, int(180 * val / mx))
            tk.Frame(trk, bg=color, height=14, width=fw).place(x=0, y=0)

            tk.Label(rf, text=str(val), bg=C_BLANCO,
                     fg=C_TEXTO, font=F_SMALL).pack(side="left")

        tk.Frame(c, bg=C_BLANCO, height=10).pack()
        return c

    # ══════════════════════════════════════════════════════════════
    #  NUEVO REGISTRO  (tu guardar_en_mysql integrada aqui)
    # ══════════════════════════════════════════════════════════════
    def _nuevo(self):
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Nuevo registro de prospecto",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer, text="Registra manualmente un alumno prospecto en la BD",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Formulario centrado
        wrap = tk.Frame(outer, bg=C_GRIS_BG)
        wrap.pack(expand=True)

        form = card_frame(wrap)
        form.pack(ipadx=6)
        tk.Frame(form, bg=C_VINO, height=4).pack(fill="x")

        tk.Label(form, text="Datos del prospecto",
                 bg=C_BLANCO, fg=C_VINO, font=F_H3).pack(
            anchor="w", padx=32, pady=(18, 12))
        tk.Frame(form, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        self._nf = {}

        def campo(etiqueta, key):
            row = tk.Frame(form, bg=C_BLANCO)
            row.pack(fill="x", padx=32, pady=8)
            tk.Label(row, text=etiqueta, bg=C_BLANCO, fg=C_TEXTO,
                     font=F_BOLD, width=24, anchor="w").pack(side="left")
            e = tk.Entry(row, font=F_BODY, bg="#F5F5F5",
                         fg=C_TEXTO, relief="flat", bd=0,
                         insertbackground=C_VINO, width=36)
            e.pack(side="left", ipady=7, fill="x", expand=True)
            self._nf[key] = e

        campo("Nombre completo  *",       "nombre")
        campo("Correo electrónico  *",    "correo")
        campo("Bachillerato de origen  *","prepa")

        # Dropdown carrera
        rc = tk.Frame(form, bg=C_BLANCO)
        rc.pack(fill="x", padx=32, pady=8)
        tk.Label(rc, text="Carrera de interés  *",
                 bg=C_BLANCO, fg=C_TEXTO, font=F_BOLD,
                 width=24, anchor="w").pack(side="left")
        self._cvar = tk.StringVar(value="aún no lo sé")
        ttk.Combobox(rc, textvariable=self._cvar,
                     values=list(DICCIONARIO_CARRERAS.keys()),
                     state="readonly", font=F_BODY, width=34).pack(
            side="left", ipady=5)

        tk.Frame(form, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        # Botones
        brow = tk.Frame(form, bg=C_BLANCO)
        brow.pack(fill="x", padx=32, pady=18)
        btn_pill(brow, "  💾  Guardar prospecto  ",
                 self._guardar, py=9).pack(side="left")
        btn_pill(brow, "  ✕ Limpiar  ",
                 self._limpiar, bg="#BDBDBD", fg=C_TEXTO, py=9).pack(
            side="left", padx=10)

        self._log = tk.StringVar(value="")
        tk.Label(form, textvariable=self._log, bg=C_BLANCO,
                 fg=C_VERDE_TAG, font=F_BOLD).pack(pady=(0, 14))

    def _guardar(self):
        datos = {k: e.get().strip() for k, e in self._nf.items()}
        datos["carrera"] = self._cvar.get()
        if not datos["nombre"] or not datos["correo"] or not datos["prepa"]:
            messagebox.showwarning("Campos requeridos",
                                   "Nombre, correo y bachillerato son obligatorios.",
                                   parent=self)
            return
        ok = self.db.guardar_prospecto(datos)
        if ok:
            self._log.set(f"✅  '{datos['nombre']}' guardado correctamente.")
            self._limpiar(keep_log=True)
        else:
            self._log.set("❌  No se pudo guardar. Revisa la consola.")

    def _limpiar(self, keep_log=False):
        for e in self._nf.values():
            e.delete(0, "end")
        self._cvar.set("aún no lo sé")
        if not keep_log:
            self._log.set("")

    # ══════════════════════════════════════════════════════════════
    #  TABLA GENERICA
    # ══════════════════════════════════════════════════════════════
    _TAG_COLORS = {
        "COBAEP":  C_AZUL_TAG,
        "CONALEP": C_NARANJO,
        "CECyTE":  C_VERDE_TAG,
        "CBTIS":   C_MORADO_TAG,
        "CBTA":    "#00695C",
    }

    def _tag_color(self, txt):
        for k, c in self._TAG_COLORS.items():
            if k.lower() in str(txt).lower():
                return c
        return "#546E7A"

    COLS = [
        ("nombre_completo", "Nombre completo", 20),
        ("correo", "Correo", 26),
        ("nombre_bachillerato", "Bachillerato", 14),
        ("nombre_carrera", "Carrera de interés", 26),
        ("estatus_proceso", "Estatus de Admisión", 10),
        ("fecha_registro", "Fecha", 8),
    ]

    def _tabla(self, parent, rows, num=False):
        if not rows:
            tk.Label(parent, text="No hay registros disponibles.",
                     bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(pady=24)
            return

        wrap = tk.Frame(parent, bg=C_GRIS_BG,
                        highlightbackground=C_GRIS_BORDE,
                        highlightthickness=1)
        wrap.pack(fill="both", expand=True)

        cv = tk.Canvas(wrap, bg=C_BLANCO, highlightthickness=0, bd=0)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=cv.yview)
        inn = tk.Frame(cv, bg=C_BLANCO)

        inn.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))

        # Guardamos la referencia de la ventana del canvas para que se estire al ancho total
        cv_window = cv.create_window((0, 0), window=inn, anchor="nw")
        cv.bind('<Configure>', lambda e: cv.itemconfig(cv_window, width=e.width))

        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        cv.bind_all("<MouseWheel>", lambda e: cv.yview_scroll(-1 * (e.delta // 120), "units"))

        # --- ENCABEZADO ---
        hdr = tk.Frame(inn, bg=C_GRIS_HDR)
        hdr.pack(fill="x")

        col_offset = 0  # Usamos esto por si activas el contador de números '#'
        if num:
            tk.Label(hdr, text="#", bg=C_GRIS_HDR, fg=C_SUBTEXTO, font=F_BOLD, anchor="center", pady=8).grid(row=0,
                                                                                                             column=0,
                                                                                                             sticky="ew",
                                                                                                             padx=8)
            hdr.grid_columnconfigure(0, weight=0, minsize=40)  # El # no se estira, se queda de tamaño fijo
            col_offset = 1

        for idx, (_, lbl, w) in enumerate(self.COLS):
            col_idx = idx + col_offset
            tk.Label(hdr, text=lbl, bg=C_GRIS_HDR, fg=C_SUBTEXTO, font=F_BOLD, anchor="w", pady=8).grid(row=0,
                                                                                                        column=col_idx,
                                                                                                        sticky="ew",
                                                                                                        padx=8)
            hdr.grid_columnconfigure(col_idx, weight=w)  # Le damos el "peso" para que se estire proporcionalmente

        tk.Frame(inn, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        # --- FILAS ---
        for i, row in enumerate(rows):
            bg = C_BLANCO if i % 2 == 0 else C_GRIS_ROW
            rf = tk.Frame(inn, bg=bg)
            rf.pack(fill="x")

            col_offset = 0
            if num:
                tk.Label(rf, text=str(i + 1), bg=bg, fg=C_SUBTEXTO, font=F_SMALL, anchor="center", pady=6).grid(row=0,
                                                                                                                column=0,
                                                                                                                sticky="ew",
                                                                                                                padx=8)
                rf.grid_columnconfigure(0, weight=0, minsize=40)
                col_offset = 1

            for idx, (key, _, w) in enumerate(self.COLS):
                col_idx = idx + col_offset
                raw = row.get(key, "")
                val = str(raw) if raw is not None else "—"

                if isinstance(raw, (datetime.datetime, datetime.date)):
                    val = raw.strftime("%d/%m/%Y")

                # Corregido: Ahora usamos la llave que declaraste en tu lista COLS ("nombre_bachillerato")
                if key == "nombre_bachillerato":
                    cf = tk.Frame(rf, bg=bg)
                    cf.grid(row=0, column=col_idx, sticky="w", padx=8, pady=4)
                    tag_pill(cf, val, self._tag_color(val)).pack(anchor="w")
                else:
                    tk.Label(rf, text=val, bg=bg, fg=C_TEXTO, font=F_SMALL, anchor="w", pady=6).grid(row=0,
                                                                                                     column=col_idx,
                                                                                                     sticky="ew",
                                                                                                     padx=8)

                rf.grid_columnconfigure(col_idx, weight=w)  # Hacemos que la fila se estire igual que el encabezado

            tk.Frame(inn, bg=C_GRIS_BORDE, height=1).pack(fill="x")

    # ══════════════════════════════════════════════════════════════
    #  EXPORTAR EXCEL
    # ══════════════════════════════════════════════════════════════
    def _exportar(self):
        fid  = self._fmap.get(self._cv.get()) if hasattr(self, "_fmap") else None
        rows = self.db.get_prospectos(fid)
        if not rows:
            messagebox.showinfo("Sin datos", "No hay prospectos para exportar.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"prospectos_{datetime.date.today()}.xlsx",
            title="Guardar lista de prospectos",
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Prospectos"

        R_FILL = PatternFill("solid", fgColor="7A0000")
        H_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        BORDE  = Border(bottom=Side(style="thin", color="E0E0E0"))
        CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        IZQ    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        ws.merge_cells("A1:G1")
        ws["A1"] = "Sistema Lince  —  Lista de Prospectos"
        ws["A1"].font      = Font(name="Georgia", size=15, bold=True, color="7A0000")
        ws["A1"].alignment = CTR
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = (f"Generado: {datetime.datetime.now():%d/%m/%Y %H:%M}"
                    f"   |   Total: {len(rows)} registros")
        ws["A2"].font      = Font(size=9, color="757575", name="Calibri")
        ws["A2"].alignment = CTR

        HDRS = ["#","Nombre completo","Correo",
                "Bachillerato de origen","Carrera de interés",
                "ID Carrera","Fecha de registro"]
        KEYS = ["_i","nombre_completo","correo",
                "bachillerato_origen","nombre_carrera",
                "id_carrera_interes","fecha_registro"]
        WCOL = [4, 28, 28, 22, 22, 10, 16]

        for ci, h in enumerate(HDRS, 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.fill = R_FILL; c.font = H_FONT
            c.alignment = CTR; c.border = BORDE
        ws.row_dimensions[4].height = 20

        for ri, row in enumerate(rows, 5):
            alt = PatternFill("solid", fgColor="FFFFFF" if ri%2 else "F9F9F9")
            for ci, key in enumerate(KEYS, 1):
                val = ri-4 if key == "_i" else row.get(key, "")
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = val.strftime("%d/%m/%Y")
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = alt
                c.alignment = CTR if ci == 1 else IZQ
                c.border = BORDE
            ws.row_dimensions[ri].height = 15

        for i, w in enumerate(WCOL, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A5"
        wb.save(path)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")

    def _exp_rep(self, tipo):
        st   = self.db.get_stats()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{tipo}_{datetime.date.today()}.xlsx",
            title=f"Guardar reporte de {tipo}",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.title()

        RF = PatternFill("solid", fgColor="7A0000")
        HF = Font(color="FFFFFF", bold=True, name="Calibri")
        CT = Alignment(horizontal="center", vertical="center")

        ws["A1"] = f"Sistema Lince — Reporte de {tipo.title()}"
        ws["A1"].font = Font(name="Georgia", size=13, bold=True, color="7A0000")

        if tipo == "bachilleratos":
            data = st.get("bach_stats", [])
            hdrs = ["Bachillerato", "Prospectos"]
            keys = ["bach", "cnt"]
        else:
            data = st.get("carrera_stats", [])
            hdrs = ["Carrera de interés", "Prospectos"]
            keys = ["nombre_carrera", "cnt"]

        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.fill = RF; c.font = HF; c.alignment = CT

        for ri, row in enumerate(data, 4):
            for ci, k in enumerate(keys, 1):
                ws.cell(row=ri, column=ci, value=row.get(k, ""))

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        wb.save(path)
        messagebox.showinfo("Reporte guardado", f"Guardado en:\n{path}")


