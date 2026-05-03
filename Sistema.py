"""
Sistema Lince — Aplicación de Escritorio
Tecnológico de Atlixco (ITSA)

Requisitos:
    pip install mysql-connector-python openpyxl

Tabla principal: prospectos
Columnas: nombre_completo, correo, bachillerato_origen,
          id_carrera_interes, fecha_registro
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
from mysql.connector import Error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import hashlib


# ══════════════════════════════════════════════════════════════════
#  PALETA DE COLORES Y TIPOGRAFIA
# ══════════════════════════════════════════════════════════════════
C_VINO       = "#7A0000"
C_ROJO_BTN   = "#B71C1C"
C_ROJO_HOV   = "#C62828"
C_BLANCO     = "#FFFFFF"
C_GRIS_BG    = "#F5F5F5"
C_GRIS_CARD  = "#FFFFFF"
C_GRIS_HDR   = "#FAFAFA"
C_GRIS_BORDE = "#E0E0E0"
C_GRIS_ROW   = "#F9F9F9"
C_TEXTO      = "#212121"
C_SUBTEXTO   = "#757575"
C_AZUL_TAG   = "#1565C0"
C_VERDE_TAG  = "#2E7D32"
C_NARANJO    = "#E65100"
C_MORADO_TAG = "#6A1B9A"

F_TITULO = ("Georgia",    22, "bold")
F_NAV    = ("Helvetica",  11, "bold")
F_H2     = ("Helvetica",  15, "bold")
F_H3     = ("Helvetica",  12, "bold")
F_BODY   = ("Helvetica",  10)
F_BOLD   = ("Helvetica",  10, "bold")
F_SMALL  = ("Helvetica",   9)
F_STAT_N = ("Helvetica",  26, "bold")
F_STAT_L = ("Helvetica",   8)


# ══════════════════════════════════════════════════════════════════
#  DICCIONARIO CARRERAS  (igual que tu funcion original)
# ══════════════════════════════════════════════════════════════════
DICCIONARIO_CARRERAS = {
    "sistemas":                            1,
    "bioquimica":                          3,
    "bioquimica":                          3,
    "mecatronica":                         4,
    "industrial":                          5,
    "electromecanica":                     2,
    "gastronomia":                         6,
    "maestria en ingenieria":              7,
    "maestria en inteligencia artificial": 8,
    "aun no lo se":                        None,
    # Con acento tambien
    "bioquímica":                          3,
    "mecatrónica":                         4,
    "electromecánica":                     2,
    "gastronomía":                         6,
    "aún no lo sé":                        None,
}

CARRERAS_NOMBRE = {
    1: "Ing. en Sistemas",
    2: "Electromecánica",
    3: "Bioquímica",
    4: "Mecatrónica",
    5: "Ing. Industrial",
    6: "Gastronomía",
    7: "Maestría en Ingeniería",
    8: "Maestría en IA",
    None: "Por definir",
}


# ══════════════════════════════════════════════════════════════════
#  CAPA DE BASE DE DATOS
# ══════════════════════════════════════════════════════════════════
class Database:
    def __init__(self):
        self.conn = None
        self._connect()

    def _connect(self):
        try:
            self.conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="blaky2507",
                database="lincebot",
                port=3307,
                connection_timeout=6,
            )
        except Error as e:
            messagebox.showerror(
                "Error de conexion",
                f"No se pudo conectar a la base de datos:\n\n{e}\n\n"
                "Verifica que MySQL este activo en el puerto 3307.",
            )

    def _ok(self) -> bool:
        try:
            if self.conn and self.conn.is_connected():
                return True
            self._connect()
            return self.conn is not None
        except Exception:
            return False

    # ── Login / registro de usuarios ──────────────────────────────
    def login(self, correo: str, password: str):
        if not self._ok():
            return None
        try:
            cur = self.conn.cursor(dictionary=True)
            cur.execute("SELECT * FROM usuarios WHERE correo = %s LIMIT 1", (correo,))
            user = cur.fetchone()
            cur.close()
            if not user:
                return None
            stored = str(user.get("password",
                          user.get("contrasena",
                          user.get("contraseña", ""))))
            if stored in (
                password,
                hashlib.md5(password.encode()).hexdigest(),
                hashlib.sha256(password.encode()).hexdigest(),
            ):
                return user
            return None
        except Exception as e:
            messagebox.showerror("Error login", str(e))
            return None

    def register(self, nombre, rol, num_trabajador, correo, password) -> bool:
        if not self._ok():
            return False
        try:
            cur = self.conn.cursor()
            cur.execute(
                "INSERT INTO usuarios "
                "(nombre, rol, numero_trabajador, correo, password) "
                "VALUES (%s, %s, %s, %s, %s)",
                (nombre, rol, num_trabajador, correo, password),
            )
            self.conn.commit()
            cur.close()
            return True
        except Exception as e:
            messagebox.showerror("Error al registrar", str(e))
            return False

    # ── guardar_en_mysql  (tu funcion original, integrada) ─────────
    def guardar_prospecto(self, datos: dict) -> bool:
        """
        Inserta en prospectos.
        datos: nombre, correo, prepa, carrera
        """
        if not self._ok():
            return False
        try:
            fecha_actual  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            carrera_texto = datos.get("carrera", "").lower()
            id_carrera    = DICCIONARIO_CARRERAS.get(carrera_texto, None)

            sql = (
                "INSERT INTO prospectos "
                "(nombre_completo, correo, bachillerato_origen, "
                " id_carrera_interes, fecha_registro) "
                "VALUES (%s, %s, %s, %s, %s)"
            )
            valores = (
                datos.get("nombre"),
                datos.get("correo"),
                datos.get("prepa"),
                id_carrera,
                fecha_actual,
            )
            cur = self.conn.cursor()
            cur.execute(sql, valores)
            self.conn.commit()
            cur.close()
            print(f"Prospecto '{datos.get('nombre')}' guardado. id_carrera={id_carrera}")
            return True
        except Exception as e:
            print(f"Error al guardar prospecto: {e}")
            messagebox.showerror("Error", f"No se pudo guardar el prospecto:\n{e}")
            return False

    # ── Consultas ──────────────────────────────────────────────────
    def get_prospectos(self, filtro_carrera_id=None):
        if not self._ok():
            return []
        try:
            cur = self.conn.cursor(dictionary=True)
            if filtro_carrera_id is not None:
                cur.execute(
                    "SELECT p.*, c.nombre_carrera "
                    "FROM prospectos p "
                    "LEFT JOIN carreras c ON p.id_carrera_interes = c.id "
                    "WHERE p.id_carrera_interes = %s "
                    "ORDER BY p.fecha_registro DESC",
                    (filtro_carrera_id,),
                )
            else:
                cur.execute(
                    "SELECT p.*, c.nombre_carrera "
                    "FROM prospectos p "
                    "LEFT JOIN carreras c ON p.id_carrera_interes = c.id "
                    "ORDER BY p.fecha_registro DESC"
                )
            rows = cur.fetchall()
            cur.close()
        except Exception:
            # Fallback sin JOIN si no existe tabla carreras
            try:
                cur = self.conn.cursor(dictionary=True)
                if filtro_carrera_id is not None:
                    cur.execute(
                        "SELECT * FROM prospectos "
                        "WHERE id_carrera_interes = %s "
                        "ORDER BY fecha_registro DESC",
                        (filtro_carrera_id,),
                    )
                else:
                    cur.execute("SELECT * FROM prospectos ORDER BY fecha_registro DESC")
                rows = cur.fetchall()
                cur.close()
            except Exception as e2:
                messagebox.showerror("Error consulta", str(e2))
                return []

        for r in rows:
            if not r.get("nombre_carrera"):
                r["nombre_carrera"] = CARRERAS_NOMBRE.get(
                    r.get("id_carrera_interes"), "—"
                )
        return rows

    def get_stats(self) -> dict:
        if not self._ok():
            return {}
        stats = {}
        try:
            cur = self.conn.cursor(dictionary=True)
            now = datetime.datetime.now()

            cur.execute("SELECT COUNT(*) AS t FROM prospectos")
            stats["total"] = cur.fetchone()["t"]

            cur.execute(
                "SELECT COUNT(*) AS m FROM prospectos "
                "WHERE MONTH(fecha_registro)=%s AND YEAR(fecha_registro)=%s",
                (now.month, now.year),
            )
            stats["mes"] = cur.fetchone()["m"]

            cur.execute(
                "SELECT COUNT(DISTINCT bachillerato_origen) AS b FROM prospectos"
            )
            stats["bachilleratos"] = cur.fetchone()["b"]

            cur.execute(
                "SELECT id_carrera_interes AS cid, COUNT(*) AS cnt "
                "FROM prospectos GROUP BY id_carrera_interes ORDER BY cnt DESC LIMIT 1"
            )
            row = cur.fetchone()
            stats["carrera_top"] = CARRERAS_NOMBRE.get(row["cid"], "—") if row else "—"
            stats["carrera_pct"] = (
                round(row["cnt"] / stats["total"] * 100)
                if row and stats["total"] else 0
            )

            cur.execute(
                "SELECT bachillerato_origen AS bach, COUNT(*) AS cnt "
                "FROM prospectos GROUP BY bachillerato_origen "
                "ORDER BY cnt DESC LIMIT 6"
            )
            stats["bach_stats"] = cur.fetchall()

            cur.execute(
                "SELECT id_carrera_interes AS cid, COUNT(*) AS cnt "
                "FROM prospectos GROUP BY id_carrera_interes "
                "ORDER BY cnt DESC LIMIT 6"
            )
            stats["carrera_stats"] = [
                {"nombre_carrera": CARRERAS_NOMBRE.get(r["cid"], "—"),
                 "cnt": r["cnt"]}
                for r in cur.fetchall()
            ]
            cur.close()
        except Exception as e:
            messagebox.showerror("Error estadisticas", str(e))
        return stats


# ══════════════════════════════════════════════════════════════════
#  HELPERS UI
# ══════════════════════════════════════════════════════════════════
def btn_pill(parent, text, cmd, bg=C_ROJO_BTN, fg=C_BLANCO,
             font=F_BOLD, px=20, py=7):
    b = tk.Button(
        parent, text=text, command=cmd,
        bg=bg, fg=fg, font=font,
        relief="flat", bd=0, cursor="hand2",
        activebackground=C_ROJO_HOV, activeforeground=C_BLANCO,
        padx=px, pady=py,
    )
    b.bind("<Enter>", lambda e, _bg=bg: b.config(bg=C_ROJO_HOV if _bg == C_ROJO_BTN else _bg))
    b.bind("<Leave>", lambda e, _bg=bg: b.config(bg=_bg))
    return b


def card_frame(parent):
    return tk.Frame(parent, bg=C_GRIS_CARD,
                    highlightbackground=C_GRIS_BORDE,
                    highlightthickness=1)


def h_sep(parent):
    tk.Frame(parent, bg=C_GRIS_BORDE, height=1).pack(fill="x")


def tag_pill(parent, text, bg):
    return tk.Label(parent, text=text, bg=bg, fg=C_BLANCO,
                    font=F_SMALL, padx=6, pady=2)


# ══════════════════════════════════════════════════════════════════
#  VENTANA LOGIN
# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ══════════════════════════════════════════════════════════════════
class MainWin(tk.Toplevel):
    TABS = ["Panel general", "Lista de alumnos", "Estadísticas", "Nuevo registro"]

    def __init__(self, master, db, user):
        super().__init__(master)
        self.db, self.user = db, user
        self.title("Sistema Lince")
        self.geometry("1140x720")
        self.minsize(920, 600)
        self.configure(bg=C_GRIS_BG)
        self._shell()
        self._nav("Panel general")
        self.grab_set()

    # ── Shell (header + area) ──────────────────────────────────────
    def _shell(self):
        hdr = tk.Frame(self, bg=C_VINO, height=86)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        inner = tk.Frame(hdr, bg=C_VINO)
        inner.pack(fill="both", expand=True, padx=22)

        # Izquierda: titulo + usuario
        left = tk.Frame(inner, bg=C_VINO)
        left.pack(side="left", pady=8)
        tk.Label(left, text="◈  Sistema Lince",
                 bg=C_VINO, fg=C_BLANCO,
                 font=("Georgia", 19, "bold")).pack(anchor="w")
        nombre_u = self.user.get("nombre", self.user.get("correo", "Usuario"))
        rol_u    = self.user.get("rol", "")
        tk.Label(left, text=f"Bienvenido, {nombre_u}  ·  {rol_u}",
                 bg=C_VINO, fg="#FF8A80", font=F_SMALL).pack(anchor="w")

        # Derecha: botones nav
        nav = tk.Frame(inner, bg=C_VINO)
        nav.pack(side="right", pady=18)
        ICONS = {"Panel general": "⊞", "Lista de alumnos": "☰",
                 "Estadísticas": "⊕", "Nuevo registro": "＋"}
        self._nb = {}
        for t in self.TABS:
            icon = ICONS.get(t, "")
            b = tk.Button(
                nav, text=f"  {icon} {t}  ", font=F_NAV,
                bg=C_VINO, fg=C_BLANCO, relief="flat",
                bd=0, cursor="hand2", padx=10, pady=5,
                activebackground="#5C0000",
                command=lambda tab=t: self._nav(tab),
            )
            b.pack(side="left", padx=2)
            self._nb[t] = b

        # Barra color separador
        tk.Frame(self, bg="#5C0000", height=3).pack(fill="x")

        self._area = tk.Frame(self, bg=C_GRIS_BG)
        self._area.pack(fill="both", expand=True)

    def _nav(self, tab):
        for t, b in self._nb.items():
            b.config(bg="#5C0000" if t == tab else C_VINO)
        for w in self._area.winfo_children():
            w.destroy()
        {
            "Panel general":    self._panel,
            "Lista de alumnos": self._lista,
            "Estadísticas":     self._stats,
            "Nuevo registro":   self._nuevo,
        }[tab]()

    # ══════════════════════════════════════════════════════════════
    #  PANEL GENERAL
    # ══════════════════════════════════════════════════════════════
    def _panel(self):
        st   = self.db.get_stats()
        now  = datetime.datetime.now()
        MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Panel general",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer,
                 text="Resumen de prospectos registrados vía chatbot",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Tarjetas estadisticas
        crow = tk.Frame(outer, bg=C_GRIS_BG)
        crow.pack(fill="x", pady=(0, 20))

        cards = [
            ("Total\nregistrados",     str(st.get("total", 0)),          C_VINO),
            (MESES[now.month]+f"\n{now.year}",
             str(st.get("mes", 0)),                                       C_AZUL_TAG),
            ("Bachilleratos\ndistintos",str(st.get("bachilleratos", 0)), C_VERDE_TAG),
            ("Carrera top\n"+st.get("carrera_top","—"),
             f"{st.get('carrera_pct',0)}%",                              C_MORADO_TAG),
        ]
        for lbl, val, accent in cards:
            c = tk.Frame(crow, bg=C_BLANCO,
                         highlightbackground=C_GRIS_BORDE,
                         highlightthickness=1)
            c.pack(side="left", fill="both", expand=True, padx=5)
            tk.Frame(c, bg=accent, height=4).pack(fill="x")
            tk.Label(c, text=val, bg=C_BLANCO, fg=accent,
                     font=F_STAT_N).pack(pady=(14, 4))
            tk.Label(c, text=lbl, bg=C_BLANCO, fg=C_SUBTEXTO,
                     font=F_STAT_L, justify="center").pack(pady=(0, 14))

        # Recientes
        rh = tk.Frame(outer, bg=C_GRIS_BG)
        rh.pack(fill="x", pady=(0, 8))
        tk.Label(rh, text="Registros recientes",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H3).pack(side="left")
        btn_pill(rh, "Ver todos →",
                 lambda: self._nav("Lista de alumnos"),
                 bg=C_BLANCO, fg=C_VINO, px=12, py=4).pack(side="right")

        self._tabla(outer, self.db.get_prospectos()[:10])

    # ══════════════════════════════════════════════════════════════
    #  LISTA DE ALUMNOS
    # ══════════════════════════════════════════════════════════════
    def _lista(self):
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Lista de alumnos",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer, text="Prospectos registrados desde el chatbot",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Controles
        ctrl = tk.Frame(outer, bg=C_GRIS_BG)
        ctrl.pack(fill="x", pady=(0, 10))

        opciones  = [("Todas las carreras", None)] + \
                    [(n, i) for i, n in CARRERAS_NOMBRE.items()]
        self._fmap = {o[0]: o[1] for o in opciones}
        nombres   = [o[0] for o in opciones]

        self._cv = tk.StringVar(value="Todas las carreras")
        ttk.Combobox(ctrl, textvariable=self._cv, values=nombres,
                     state="readonly", font=F_BODY, width=28).pack(
            side="left", ipady=4, padx=(0, 10))
        self._cv.trace_add("write", lambda *_: self._reload_tabla())

        btn_pill(ctrl, "⬇  Exportar Excel", self._exportar,
                 bg="#37474F", px=14, py=5).pack(side="left")

        self._clbl = tk.Label(ctrl, text="", bg=C_GRIS_BG,
                              fg=C_SUBTEXTO, font=F_SMALL)
        self._clbl.pack(side="right")

        self._touter = tk.Frame(outer, bg=C_GRIS_BG)
        self._touter.pack(fill="both", expand=True)
        self._reload_tabla()

    def _reload_tabla(self):
        for w in self._touter.winfo_children():
            w.destroy()
        fid  = self._fmap.get(self._cv.get()) if hasattr(self, "_cv") else None
        rows = self.db.get_prospectos(fid)
        self._clbl.config(text=f"{len(rows)} registros")
        self._tabla(self._touter, rows, num=True)

    # ══════════════════════════════════════════════════════════════
    #  ESTADISTICAS
    # ══════════════════════════════════════════════════════════════
    def _stats(self):
        st    = self.db.get_stats()
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Estadísticas",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer,
                 text="Bachilleratos con más registros y carreras de mayor interés",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        brow = tk.Frame(outer, bg=C_GRIS_BG)
        brow.pack(fill="x", pady=(0, 18))
        btn_pill(brow, "⬇  Reporte bachilleratos",
                 lambda: self._exp_rep("bachilleratos"),
                 bg="#37474F", px=14, py=5).pack(side="left", padx=(0, 10))
        btn_pill(brow, "⬇  Reporte carreras",
                 lambda: self._exp_rep("carreras"),
                 bg="#37474F", px=14, py=5).pack(side="left")

        gr = tk.Frame(outer, bg=C_GRIS_BG)
        gr.pack(fill="both", expand=True)

        self._barchart(
            gr, "Bachilleratos con más alumnos",
            st.get("bach_stats", []),
            "bach", "cnt", C_AZUL_TAG,
        ).pack(side="left", fill="both", expand=True, padx=(0, 12))

        self._barchart(
            gr, "Carreras con mayor interés",
            st.get("carrera_stats", []),
            "nombre_carrera", "cnt", C_VERDE_TAG,
        ).pack(side="left", fill="both", expand=True)

    def _barchart(self, parent, title, data, lkey, vkey, color):
        c = card_frame(parent)
        tk.Frame(c, bg=color, height=4).pack(fill="x")
        tk.Label(c, text=title, bg=C_BLANCO, fg=C_TEXTO,
                 font=F_H3, anchor="w").pack(anchor="w", padx=16, pady=(12, 8))
        tk.Frame(c, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        mx = max((r[vkey] for r in data), default=1) or 1
        for r in data:
            lbl = str(r.get(lkey, "—"))
            val = r.get(vkey, 0)
            rf  = tk.Frame(c, bg=C_BLANCO)
            rf.pack(fill="x", padx=16, pady=5)

            tk.Label(rf, text=lbl, bg=C_BLANCO, fg=C_TEXTO,
                     font=F_SMALL, width=20, anchor="w").pack(side="left")

            trk = tk.Frame(rf, bg="#EEEEEE", height=14, width=180)
            trk.pack(side="left", padx=6)
            trk.pack_propagate(False)
            fw = max(2, int(180 * val / mx))
            tk.Frame(trk, bg=color, height=14, width=fw).place(x=0, y=0)

            tk.Label(rf, text=str(val), bg=C_BLANCO,
                     fg=C_TEXTO, font=F_SMALL).pack(side="left")

        tk.Frame(c, bg=C_BLANCO, height=10).pack()
        return c

    # ══════════════════════════════════════════════════════════════
    #  NUEVO REGISTRO  (tu guardar_en_mysql integrada aqui)
    # ══════════════════════════════════════════════════════════════
    def _nuevo(self):
        outer = tk.Frame(self._area, bg=C_GRIS_BG)
        outer.pack(fill="both", expand=True, padx=28, pady=20)

        tk.Label(outer, text="Nuevo registro de prospecto",
                 bg=C_GRIS_BG, fg=C_TEXTO, font=F_H2).pack(anchor="w")
        tk.Label(outer, text="Registra manualmente un alumno prospecto en la BD",
                 bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(anchor="w")
        tk.Frame(outer, bg=C_GRIS_BORDE, height=1).pack(fill="x", pady=12)

        # Formulario centrado
        wrap = tk.Frame(outer, bg=C_GRIS_BG)
        wrap.pack(expand=True)

        form = card_frame(wrap)
        form.pack(ipadx=6)
        tk.Frame(form, bg=C_VINO, height=4).pack(fill="x")

        tk.Label(form, text="Datos del prospecto",
                 bg=C_BLANCO, fg=C_VINO, font=F_H3).pack(
            anchor="w", padx=32, pady=(18, 12))
        tk.Frame(form, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        self._nf = {}

        def campo(etiqueta, key):
            row = tk.Frame(form, bg=C_BLANCO)
            row.pack(fill="x", padx=32, pady=8)
            tk.Label(row, text=etiqueta, bg=C_BLANCO, fg=C_TEXTO,
                     font=F_BOLD, width=24, anchor="w").pack(side="left")
            e = tk.Entry(row, font=F_BODY, bg="#F5F5F5",
                         fg=C_TEXTO, relief="flat", bd=0,
                         insertbackground=C_VINO, width=36)
            e.pack(side="left", ipady=7, fill="x", expand=True)
            self._nf[key] = e

        campo("Nombre completo  *",       "nombre")
        campo("Correo electrónico  *",    "correo")
        campo("Bachillerato de origen  *","prepa")

        # Dropdown carrera
        rc = tk.Frame(form, bg=C_BLANCO)
        rc.pack(fill="x", padx=32, pady=8)
        tk.Label(rc, text="Carrera de interés  *",
                 bg=C_BLANCO, fg=C_TEXTO, font=F_BOLD,
                 width=24, anchor="w").pack(side="left")
        self._cvar = tk.StringVar(value="aún no lo sé")
        ttk.Combobox(rc, textvariable=self._cvar,
                     values=list(DICCIONARIO_CARRERAS.keys()),
                     state="readonly", font=F_BODY, width=34).pack(
            side="left", ipady=5)

        tk.Frame(form, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        # Botones
        brow = tk.Frame(form, bg=C_BLANCO)
        brow.pack(fill="x", padx=32, pady=18)
        btn_pill(brow, "  💾  Guardar prospecto  ",
                 self._guardar, py=9).pack(side="left")
        btn_pill(brow, "  ✕ Limpiar  ",
                 self._limpiar, bg="#BDBDBD", fg=C_TEXTO, py=9).pack(
            side="left", padx=10)

        self._log = tk.StringVar(value="")
        tk.Label(form, textvariable=self._log, bg=C_BLANCO,
                 fg=C_VERDE_TAG, font=F_BOLD).pack(pady=(0, 14))

    def _guardar(self):
        datos = {k: e.get().strip() for k, e in self._nf.items()}
        datos["carrera"] = self._cvar.get()
        if not datos["nombre"] or not datos["correo"] or not datos["prepa"]:
            messagebox.showwarning("Campos requeridos",
                                   "Nombre, correo y bachillerato son obligatorios.",
                                   parent=self)
            return
        ok = self.db.guardar_prospecto(datos)
        if ok:
            self._log.set(f"✅  '{datos['nombre']}' guardado correctamente.")
            self._limpiar(keep_log=True)
        else:
            self._log.set("❌  No se pudo guardar. Revisa la consola.")

    def _limpiar(self, keep_log=False):
        for e in self._nf.values():
            e.delete(0, "end")
        self._cvar.set("aún no lo sé")
        if not keep_log:
            self._log.set("")

    # ══════════════════════════════════════════════════════════════
    #  TABLA GENERICA
    # ══════════════════════════════════════════════════════════════
    _TAG_COLORS = {
        "COBAEP":  C_AZUL_TAG,
        "CONALEP": C_NARANJO,
        "CECyTE":  C_VERDE_TAG,
        "CBTIS":   C_MORADO_TAG,
        "CBTA":    "#00695C",
    }

    def _tag_color(self, txt):
        for k, c in self._TAG_COLORS.items():
            if k.lower() in str(txt).lower():
                return c
        return "#546E7A"

    COLS = [
        ("nombre_completo",     "Nombre completo",   22),
        ("correo",              "Correo",            22),
        ("bachillerato_origen", "Bachillerato",      14),
        ("nombre_carrera",      "Carrera de interés",18),
        ("fecha_registro",      "Fecha",             12),
    ]

    def _tabla(self, parent, rows, num=False):
        if not rows:
            tk.Label(parent, text="No hay registros disponibles.",
                     bg=C_GRIS_BG, fg=C_SUBTEXTO, font=F_BODY).pack(pady=24)
            return

        wrap = tk.Frame(parent, bg=C_GRIS_BG,
                        highlightbackground=C_GRIS_BORDE,
                        highlightthickness=1)
        wrap.pack(fill="both", expand=True)

        cv = tk.Canvas(wrap, bg=C_BLANCO, highlightthickness=0, bd=0)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=cv.yview)
        inn = tk.Frame(cv, bg=C_BLANCO)
        inn.bind("<Configure>",
                 lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=inn, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        cv.bind_all("<MouseWheel>",
                    lambda e: cv.yview_scroll(-1*(e.delta//120), "units"))

        # Encabezado
        hdr = tk.Frame(inn, bg=C_GRIS_HDR)
        hdr.pack(fill="x")
        if num:
            tk.Label(hdr, text="#", bg=C_GRIS_HDR, fg=C_SUBTEXTO,
                     font=F_BOLD, width=4, anchor="center",
                     pady=8).pack(side="left", padx=(8, 0))
        for _, lbl, w in self.COLS:
            tk.Label(hdr, text=lbl, bg=C_GRIS_HDR, fg=C_SUBTEXTO,
                     font=F_BOLD, width=w, anchor="w",
                     pady=8).pack(side="left", padx=8)
        tk.Frame(inn, bg=C_GRIS_BORDE, height=1).pack(fill="x")

        # Filas
        for i, row in enumerate(rows):
            bg = C_BLANCO if i % 2 == 0 else C_GRIS_ROW
            rf = tk.Frame(inn, bg=bg)
            rf.pack(fill="x")

            if num:
                tk.Label(rf, text=str(i+1), bg=bg, fg=C_SUBTEXTO,
                         font=F_SMALL, width=4, anchor="center",
                         pady=6).pack(side="left", padx=(8, 0))

            for key, _, w in self.COLS:
                raw = row.get(key, "")
                val = str(raw) if raw is not None else "—"
                if isinstance(raw, (datetime.datetime, datetime.date)):
                    val = raw.strftime("%d/%m/%Y")

                if key == "bachillerato_origen":
                    cf = tk.Frame(rf, bg=bg)
                    cf.pack(side="left", padx=8, pady=4)
                    tag_pill(cf, val, self._tag_color(val)).pack()
                else:
                    tk.Label(rf, text=val, bg=bg, fg=C_TEXTO,
                             font=F_SMALL, width=w, anchor="w",
                             pady=6).pack(side="left", padx=8)

            tk.Frame(inn, bg=C_GRIS_BORDE, height=1).pack(fill="x")

    # ══════════════════════════════════════════════════════════════
    #  EXPORTAR EXCEL
    # ══════════════════════════════════════════════════════════════
    def _exportar(self):
        fid  = self._fmap.get(self._cv.get()) if hasattr(self, "_fmap") else None
        rows = self.db.get_prospectos(fid)
        if not rows:
            messagebox.showinfo("Sin datos", "No hay prospectos para exportar.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"prospectos_{datetime.date.today()}.xlsx",
            title="Guardar lista de prospectos",
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Prospectos"

        R_FILL = PatternFill("solid", fgColor="7A0000")
        H_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        BORDE  = Border(bottom=Side(style="thin", color="E0E0E0"))
        CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        IZQ    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        ws.merge_cells("A1:G1")
        ws["A1"] = "Sistema Lince  —  Lista de Prospectos"
        ws["A1"].font      = Font(name="Georgia", size=15, bold=True, color="7A0000")
        ws["A1"].alignment = CTR
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = (f"Generado: {datetime.datetime.now():%d/%m/%Y %H:%M}"
                    f"   |   Total: {len(rows)} registros")
        ws["A2"].font      = Font(size=9, color="757575", name="Calibri")
        ws["A2"].alignment = CTR

        HDRS = ["#","Nombre completo","Correo",
                "Bachillerato de origen","Carrera de interés",
                "ID Carrera","Fecha de registro"]
        KEYS = ["_i","nombre_completo","correo",
                "bachillerato_origen","nombre_carrera",
                "id_carrera_interes","fecha_registro"]
        WCOL = [4, 28, 28, 22, 22, 10, 16]

        for ci, h in enumerate(HDRS, 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.fill = R_FILL; c.font = H_FONT
            c.alignment = CTR; c.border = BORDE
        ws.row_dimensions[4].height = 20

        for ri, row in enumerate(rows, 5):
            alt = PatternFill("solid", fgColor="FFFFFF" if ri%2 else "F9F9F9")
            for ci, key in enumerate(KEYS, 1):
                val = ri-4 if key == "_i" else row.get(key, "")
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = val.strftime("%d/%m/%Y")
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = alt
                c.alignment = CTR if ci == 1 else IZQ
                c.border = BORDE
            ws.row_dimensions[ri].height = 15

        for i, w in enumerate(WCOL, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A5"
        wb.save(path)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")

    def _exp_rep(self, tipo):
        st   = self.db.get_stats()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{tipo}_{datetime.date.today()}.xlsx",
            title=f"Guardar reporte de {tipo}",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.title()

        RF = PatternFill("solid", fgColor="7A0000")
        HF = Font(color="FFFFFF", bold=True, name="Calibri")
        CT = Alignment(horizontal="center", vertical="center")

        ws["A1"] = f"Sistema Lince — Reporte de {tipo.title()}"
        ws["A1"].font = Font(name="Georgia", size=13, bold=True, color="7A0000")

        if tipo == "bachilleratos":
            data = st.get("bach_stats", [])
            hdrs = ["Bachillerato", "Prospectos"]
            keys = ["bach", "cnt"]
        else:
            data = st.get("carrera_stats", [])
            hdrs = ["Carrera de interés", "Prospectos"]
            keys = ["nombre_carrera", "cnt"]

        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.fill = RF; c.font = HF; c.alignment = CT

        for ri, row in enumerate(data, 4):
            for ci, k in enumerate(keys, 1):
                ws.cell(row=ri, column=ci, value=row.get(k, ""))

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        wb.save(path)
        messagebox.showinfo("Reporte guardado", f"Guardado en:\n{path}")


# ══════════════════════════════════════════════════════════════════
#  LOGIN DIRECTO EN VENTANA RAIZ  (sin Toplevel, sin ventana en blanco)
# ══════════════════════════════════════════════════════════════════
class LoginRoot(tk.Tk):
    """
    Ventana de login construida sobre la raíz Tk.
    Evita el bug de ventana en blanco en Windows cuando se usa
    Toplevel + withdraw() sobre la raíz.
    """
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.title("Sistema Lince — Iniciar Sesión")
        self.geometry("860x520")
        self.resizable(False, False)
        self.configure(bg=C_VINO)
        # Centrar en pantalla
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - 860) // 2
        y = (self.winfo_screenheight() - 520) // 2
        self.geometry(f"860x520+{x}+{y}")
        self._ui()

    def _ui(self):
        # Fondo decorativo
        tk.Label(self, text="PROCESO ADMISIÓN",
                 bg=C_VINO, fg="#5C0000",
                 font=("Helvetica", 38, "bold")).place(x=0, y=18, width=860)
        tk.Label(self, text="PROCESO ADMISIÓN",
                 bg=C_VINO, fg="#5C0000",
                 font=("Helvetica", 38, "bold")).place(x=0, y=68, width=860)

        # Tarjeta blanca
        f = tk.Frame(self, bg=C_BLANCO)
        f.place(relx=.5, rely=.44, anchor="center", width=370, height=300)

        tk.Frame(f, bg=C_VINO, height=4).pack(fill="x")
        tk.Label(f, text="Iniciar Sesión", bg=C_BLANCO, fg=C_VINO,
                 font=("Georgia", 18, "bold")).pack(pady=(22, 16))

        self._e_correo = self._field(f, "Correo electrónico", False)
        self._e_correo.pack(fill="x", padx=30, ipady=9, pady=(0, 10))
        self._e_pass = self._field(f, "Contraseña", True)
        self._e_pass.pack(fill="x", padx=30, ipady=9, pady=(0, 18))

        btn_pill(f, "  Entrar  ", self._login, py=9).pack(
            fill="x", padx=30, pady=(0, 10))

        lnk = tk.Label(f, text="Crear cuenta", bg=C_BLANCO,
                       fg=C_SUBTEXTO, font=F_BOLD, cursor="hand2")
        lnk.pack()
        lnk.bind("<Button-1>", lambda _: self._ir_reg())
        lnk.bind("<Enter>",    lambda _: lnk.config(fg=C_VINO))
        lnk.bind("<Leave>",    lambda _: lnk.config(fg=C_SUBTEXTO))

        tk.Label(self, text="EL  ITSA  SE  TRANSFORMA",
                 bg=C_VINO, fg=C_BLANCO,
                 font=("Helvetica", 13, "bold")).place(
            relx=.5, rely=.93, anchor="center")

    def _field(self, parent, ph, secret):
        e = tk.Entry(parent, font=F_BODY, bg="#F0F0F0", fg=C_SUBTEXTO,
                     relief="flat", bd=0, insertbackground=C_VINO)
        e.insert(0, ph)
        if secret:
            e.bind("<FocusIn>",  lambda _: (e.get() == ph and (
                e.delete(0, "end"), e.config(show="*", fg=C_TEXTO))))
            e.bind("<FocusOut>", lambda _: (not e.get() and (
                e.config(show="", fg=C_SUBTEXTO), e.insert(0, ph))))
        else:
            e.bind("<FocusIn>",  lambda _: (e.get() == ph and (
                e.delete(0, "end"), e.config(fg=C_TEXTO))))
            e.bind("<FocusOut>", lambda _: (not e.get() and (
                e.config(fg=C_SUBTEXTO), e.insert(0, ph))))
        return e

    def _login(self):
        c = self._e_correo.get().strip()
        p = self._e_pass.get().strip()
        if c in ("", "Correo electrónico") or p in ("", "Contraseña"):
            messagebox.showwarning("Campos vacíos",
                                   "Ingresa correo y contraseña.", parent=self)
            return
        user = self.db.login(c, p)
        if user:
            self._abrir_main(user)
        else:
            messagebox.showerror("Acceso denegado",
                                 "Correo o contraseña incorrectos.", parent=self)

    def _ir_reg(self):
        """Oculta el login y abre registro como Toplevel."""
        self.withdraw()
        reg = RegisterRoot(self, self.db)
        reg.grab_set()

    def _abrir_main(self, user):
        """Destruye el login y abre la ventana principal."""
        self.destroy()
        root_main = tk.Tk()
        root_main.withdraw()
        _apply_style(root_main)
        app = MainWin(root_main, self.db, user)
        app.protocol("WM_DELETE_WINDOW", root_main.destroy)
        root_main.mainloop()


class RegisterRoot(tk.Toplevel):
    """
    Registro de usuario como Toplevel del LoginRoot.
    Al terminar regresa al login.
    """
    ROLES = ["Administrador", "Docente", "Coordinador", "Otro"]

    def __init__(self, login_root: LoginRoot, db: Database):
        super().__init__(login_root)
        self.login_root = login_root
        self.db = db
        self.title("Sistema Lince — Crear Cuenta")
        self.geometry("860x540")
        self.resizable(False, False)
        self.configure(bg=C_BLANCO)
        # Centrar
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - 860) // 2
        y = (self.winfo_screenheight() - 540) // 2
        self.geometry(f"860x540+{x}+{y}")
        self.protocol("WM_DELETE_WINDOW", self._volver)
        self._ui()

    def _ui(self):
        left = tk.Frame(self, bg=C_BLANCO)
        left.place(x=0, y=0, width=540, height=540)

        tk.Label(left, text="Crear Cuenta", bg=C_BLANCO, fg=C_VINO,
                 font=("Georgia", 20, "bold")).pack(pady=(36, 4))
        tk.Label(left, text="Tecnológico de Atlixco — Sistema Lince",
                 bg=C_BLANCO, fg=C_SUBTEXTO, font=F_SMALL).pack(pady=(0, 20))

        self._ef = {}
        for key, ph, sec in [
            ("nombre",         "Nombre completo",      False),
            ("num_trabajador", "Número de trabajador", False),
            ("correo",         "Correo electrónico",   False),
            ("password",       "Contraseña",           True),
            ("confirm",        "Confirmar contraseña", True),
        ]:
            e = tk.Entry(left, font=F_BODY, bg="#F0F0F0",
                         fg=C_SUBTEXTO, relief="flat", bd=0,
                         insertbackground=C_VINO,
                         show=("*" if sec else ""))
            e.insert(0, ph)
            if sec:
                e.bind("<FocusIn>",  lambda ev, en=e, p=ph: (
                    en.get() == p and (en.delete(0, "end"),
                                       en.config(show="*", fg=C_TEXTO))))
                e.bind("<FocusOut>", lambda ev, en=e, p=ph: (
                    not en.get() and (en.config(show="", fg=C_SUBTEXTO),
                                      en.insert(0, p))))
            else:
                e.bind("<FocusIn>",  lambda ev, en=e, p=ph: (
                    en.get() == p and (en.delete(0, "end"),
                                       en.config(fg=C_TEXTO))))
                e.bind("<FocusOut>", lambda ev, en=e, p=ph: (
                    not en.get() and (en.config(fg=C_SUBTEXTO),
                                      en.insert(0, p))))
            e.pack(fill="x", padx=48, ipady=9, pady=4)
            self._ef[key] = e

        self._rol = tk.StringVar(value="Selecciona un Rol")
        ttk.Combobox(left, textvariable=self._rol, values=self.ROLES,
                     state="readonly", font=F_BODY).pack(
            fill="x", padx=48, ipady=5, pady=4)

        btn_pill(left, "Registrarse", self._reg, py=10).pack(
            fill="x", padx=48, pady=(20, 0))

        lnk = tk.Label(left, text="¿Ya tienes cuenta? Inicia sesión",
                       bg=C_BLANCO, fg=C_SUBTEXTO, font=F_SMALL, cursor="hand2")
        lnk.pack(pady=8)
        lnk.bind("<Button-1>", lambda _: self._volver())

        # Panel decorativo derecho
        right = tk.Frame(self, bg=C_VINO)
        right.place(x=540, y=0, width=320, height=540)
        tk.Label(right, text="TECNOLÓGICO\nATLIXCO",
                 bg=C_VINO, fg=C_BLANCO,
                 font=("Helvetica", 24, "bold"), justify="center").place(
            relx=.5, rely=.28, anchor="center")
        tk.Label(right, text="Sistema Lince\nAdmisiones",
                 bg=C_VINO, fg="#FF8A80",
                 font=("Helvetica", 12), justify="center").place(
            relx=.5, rely=.45, anchor="center")

    def _volver(self):
        self.destroy()
        self.login_root.deiconify()

    def _reg(self):
        ph   = {"Nombre completo", "Número de trabajador",
                "Correo electrónico", "Contraseña", "Confirmar contraseña"}
        vals = {k: e.get().strip() for k, e in self._ef.items()}
        if any(v in ph or not v for v in vals.values()):
            messagebox.showwarning("Campos incompletos",
                                   "Completa todos los campos.", parent=self)
            return
        if self._rol.get() == "Selecciona un Rol":
            messagebox.showwarning("Rol requerido",
                                   "Selecciona un rol.", parent=self)
            return
        if vals["password"] != vals["confirm"]:
            messagebox.showerror("Error",
                                 "Las contraseñas no coinciden.", parent=self)
            return
        if self.db.register(vals["nombre"], self._rol.get(),
                            vals["num_trabajador"],
                            vals["correo"], vals["password"]):
            messagebox.showinfo("Cuenta creada",
                                "Registro exitoso. Ahora inicia sesión.", parent=self)
            self._volver()


def _apply_style(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("TCombobox",
                    fieldbackground="#F0F0F0",
                    background="#F0F0F0",
                    relief="flat")
    style.configure("Vertical.TScrollbar",
                    troughcolor=C_GRIS_BG,
                    background="#BDBDBD",
                    relief="flat")


if __name__ == "__main__":
    db  = Database()
    app = LoginRoot(db)
    _apply_style(app)
    app.mainloop()